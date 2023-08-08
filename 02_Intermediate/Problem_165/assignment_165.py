'''
Write a Python Program to write data to excel cell 
using python other than xlsxwriter library 
'''
# pip install openpyxl

import openpyxl

def write_to_excel(file_path, sheet_name, cell_row, cell_column, data):
    try:
        # Load the existing workbook or create a new one if it doesn't exist
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    # Select the sheet
    sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active

    # Write data to the cell
    sheet.cell(row=cell_row, column=cell_column, value=data)

    # Save the changes
    workbook.save(file_path)
    print(f"Data '{data}' written to cell {cell_row}, {cell_column} in sheet '{sheet_name}'.")

if __name__=="__main__":
    file_path = 'example.xlsx'
    sheet_name = 'Sheet1'
    cell_row = 2
    cell_column = 3
    data_to_write = "Hello, World!"

    write_to_excel(file_path, sheet_name, cell_row, cell_column, data_to_write)
