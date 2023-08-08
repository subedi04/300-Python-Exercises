'''
Write a Python Program to use 
append method in excel using python
'''

import openpyxl

def append_data_to_excel(file_path, sheet_name, data):
    try:
        # Load the existing workbook
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        # If the workbook doesn't exist, create a new one
        workbook = openpyxl.Workbook()

    # Select the sheet
    sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active

    # Append the data as a new row in the sheet
    sheet.append(data)

    # Save the changes
    workbook.save(file_path)
    print(f"Data appended to sheet '{sheet_name}'.")

if __name__=="__main__":
    file_path = 'example.xlsx'
    sheet_name = 'Sheet1'
    data_to_append = [10, 20, 30, 40, 50]

    append_data_to_excel(file_path, sheet_name, data_to_append)
