'''
Write a Python Program to perform arithmetic operation on excel file.
'''
import openpyxl

def perform_arithmetic_operation(file_path, sheet_name, operation_column, result_column):
    try:
        # Load the existing workbook
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        print(f"File '{file_path}' not found.")
        return

    # Select the sheet
    sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active

    # Perform the arithmetic operation and update the result column
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Start from the second row (skip header)
        try:
            value1 = row[operation_column - 1]
            value2 = row[operation_column]
            result = value1 + value2  # Change the operation as per your requirement
            sheet.cell(row=row[0].row, column=result_column, value=result)
        except (TypeError, ValueError):  # Handle any possible errors in case of invalid data
            continue

    workbook.save(file_path)
    print(f"Arithmetic operation performed and results written to sheet '{sheet_name}'.")

if __name__=="__main__":
    file_path = 'example.xlsx'
    sheet_name = 'Sheet1'
    operation_column = 2  # Assuming the first data column is B (column index 2)
    result_column = 3     # Assuming the result will be written in C column (column index 3)

    perform_arithmetic_operation(file_path, sheet_name, operation_column, result_column)
