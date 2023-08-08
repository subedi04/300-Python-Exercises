'''
Write a Python Program to copy data 
from one excel sheet to another
'''

import openpyxl

def copy_data_between_sheets(source_file, source_sheet_name, destination_file, destination_sheet_name):
    try:
        # Load the source workbook
        source_workbook = openpyxl.load_workbook(source_file)
    except FileNotFoundError:
        print(f"Source file '{source_file}' not found.")
        return

    # Select the source sheet
    source_sheet = source_workbook[source_sheet_name] if source_sheet_name in source_workbook.sheetnames else source_workbook.active

    # Load the destination workbook or create a new one if it doesn't exist
    try:
        destination_workbook = openpyxl.load_workbook(destination_file)
    except FileNotFoundError:
        destination_workbook = openpyxl.Workbook()

    # Select the destination sheet
    destination_sheet = destination_workbook[destination_sheet_name] if destination_sheet_name in destination_workbook.sheetnames else destination_workbook.active

    # Copy data from source to destination sheet
    for row in source_sheet.iter_rows(values_only=True):
        destination_sheet.append(row)

    # Save the changes to the destination workbook
    destination_workbook.save(destination_file)
    print(f"Data copied from '{source_sheet_name}' to '{destination_sheet_name}'.")

if __name__=="__main__":
    source_file = 'source.xlsx'
    source_sheet_name = 'Sheet1'
    destination_file = 'destination.xlsx'
    destination_sheet_name = 'Sheet2'

    copy_data_between_sheets(source_file, source_sheet_name, destination_file, destination_sheet_name)
